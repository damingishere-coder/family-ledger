from __future__ import annotations

import csv
from datetime import date, datetime
import io
import re
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .common import (
    ParsedEntry,
    ParsedSnapshot,
    decode_text,
    infer_account_type,
    month_end,
    normalize_institution,
    normalize_member_name,
    normalized_account_name,
    parse_money_to_cents,
)


ALIASES = {
    "snapshot_date": {"snapshot_date", "date", "日期", "盘点日期"},
    "member_name": {"member", "member_name", "成员", "家庭成员"},
    "account_name": {"account", "account_name", "账户", "账户名称"},
    "account_type": {"account_type", "type", "类型", "账户类型"},
    "amount": {"amount", "金额", "余额", "本期余额", "待还"},
    "amount_cents": {"amount_cents", "金额_分", "金额（分）"},
    "credit_limit": {"credit_limit", "信用额度", "额度"},
    "credit_limit_cents": {"credit_limit_cents", "信用额度_分", "信用额度（分）"},
    "billing_day": {"billing_day", "还款日", "账单日"},
    "include": {"include_in_net_worth", "include", "计入净资产", "是否计入家庭净资产"},
    "institution": {"institution", "机构", "机构名称"},
}
ACCOUNT_TYPES = {
    "wallet", "debit_card", "credit_card", "investment", "receivable",
    "other_asset", "other_liability",
}
SHEET_MONTH_RE = re.compile(r"(?:(?P<short_year>2[456])\.)?(?P<month>1[0-2]|[1-9])\s*月")
SUMMARY_ROWS = {
    "家庭存款": "total_assets_cents",
    "家庭总余额": "total_assets_cents",
    "账单": "total_liabilities_cents",
    "家庭总负债": "total_liabilities_cents",
    "家庭顺差": "net_worth_cents",
    "顺差": "net_worth_cents",
}


def _canonical_headers(headers: list[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(headers):
        normalized = str(value or "").strip().lower()
        for canonical, aliases in ALIASES.items():
            if normalized in aliases:
                result[canonical] = index
    return result


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    return date.fromisoformat(text)


def _cell(row: list[object], index: int) -> object:
    return row[index] if index < len(row) else None


def _money_value(value: object) -> tuple[int | None, list[str]]:
    if isinstance(value, float):
        value = round(value, 2)
    return parse_money_to_cents(value)


def rows_to_snapshots(rows: list[list[object]], source_sheet: str | None = None) -> list[ParsedSnapshot]:
    if not rows:
        raise ValueError("表格为空")
    headers = _canonical_headers(rows[0])
    required = {"snapshot_date", "member_name", "account_name"}
    missing = required - headers.keys()
    if "amount" not in headers and "amount_cents" not in headers:
        missing.add("amount")
    if missing:
        raise ValueError(f"表格缺少必要列：{', '.join(sorted(missing))}")

    grouped: dict[date, ParsedSnapshot] = {}
    row_errors: list[str] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            source_date = _parse_date(_cell(row, headers["snapshot_date"]))
            snapshot_date = month_end(source_date)
            member_name = normalize_member_name(str(_cell(row, headers["member_name"]) or "").strip())
            raw_name = str(_cell(row, headers["account_name"]) or "").strip()
            if not member_name or not raw_name:
                raise ValueError("成员或账户名称为空")
            if "amount" in headers:
                raw_amount = _cell(row, headers["amount"])
                amount, amount_warnings = _money_value(raw_amount)
            else:
                raw_amount = _cell(row, headers["amount_cents"])
                amount = None if raw_amount in (None, "") else int(raw_amount)
                amount_warnings = []
            category = str(_cell(row, headers["account_type"]) or "").strip() if "account_type" in headers else None
            account_type = category if category in ACCOUNT_TYPES else infer_account_type(raw_name, category)
            credit_limit = None
            if "credit_limit" in headers:
                credit_limit, limit_warnings = _money_value(_cell(row, headers["credit_limit"]))
                amount_warnings.extend(limit_warnings)
            elif "credit_limit_cents" in headers:
                raw_limit = _cell(row, headers["credit_limit_cents"])
                credit_limit = None if raw_limit in (None, "") else int(raw_limit)
            include = True
            if "include" in headers:
                raw_include = str(_cell(row, headers["include"]) or "").strip().lower()
                include = raw_include not in {"0", "false", "否", "不计入", "no"}
            institution = normalize_institution(
                str(_cell(row, headers["institution"]) or "").strip() or None
                if "institution" in headers else None
            )
            billing_day = None
            if "billing_day" in headers and _cell(row, headers["billing_day"]) not in (None, ""):
                billing_day = int(_cell(row, headers["billing_day"]))
                if not 1 <= billing_day <= 31:
                    raise ValueError("还款日必须在 1 到 31 之间")
            account_name = normalized_account_name(account_type, raw_name, institution)
            parsed = grouped.setdefault(snapshot_date, ParsedSnapshot(
                snapshot_date=snapshot_date, source_date=source_date,
                layout="flat", source_sheet=source_sheet,
            ))
            parsed.entries.append(ParsedEntry(
                member_name=member_name, account_name=account_name,
                account_type=account_type, amount_cents=amount,
                credit_limit_cents=credit_limit, include_in_net_worth=include,
                institution=institution, billing_day=billing_day,
                source_sheet=source_sheet, source_location=f"{source_sheet or '表格'}!第 {row_number} 行",
                raw_name=raw_name, raw_value="" if raw_amount is None else str(raw_amount), warnings=amount_warnings,
            ))
        except (ValueError, IndexError) as exc:
            row_errors.append(f"第 {row_number} 行：{exc}")
    if not grouped or not any(snapshot.entries for snapshot in grouped.values()):
        raise ValueError("表格中没有可导入的有效数据行")
    results = [grouped[key] for key in sorted(grouped)]
    if row_errors:
        results.append(ParsedSnapshot(
            snapshot_date=None, layout="flat", source_sheet=source_sheet,
            status="blocked", blocking_errors=row_errors,
        ))
    return results


def parse_csv(content: bytes) -> list[ParsedSnapshot]:
    snapshots, _encoding = parse_csv_with_encoding(content)
    return snapshots


def parse_csv_with_encoding(content: bytes) -> tuple[list[ParsedSnapshot], str]:
    text, encoding = decode_text(content)
    rows = [list(row) for row in csv.reader(io.StringIO(text))]
    return rows_to_snapshots(rows), encoding


def _legacy_layout(rows: list[list[object]]) -> bool:
    if len(rows) < 10:
        return False
    headers = {str(value or "").strip() for value in rows[1]}
    return {"微信", "支付宝"}.issubset(headers)


def _sheet_period(title: str, source_date: date) -> tuple[int, int] | None:
    match = SHEET_MONTH_RE.search(title)
    if not match:
        return None
    short_year = match.group("short_year")
    year = 2000 + int(short_year) if short_year else source_date.year
    return year, int(match.group("month"))


def _source_override(
    overrides: dict[object, date] | None, title: str, year: int, month: int,
) -> date | None:
    if not overrides:
        return None
    return overrides.get(title) or overrides.get((year, month))


def _formula_cache_errors(raw_sheet, cached_sheet) -> list[str]:
    errors: list[str] = []
    for row in raw_sheet.iter_rows():
        for raw_cell in row:
            if raw_cell.data_type == "f":
                cached = cached_sheet[raw_cell.coordinate].value
                if cached is None or (isinstance(cached, str) and not cached.strip()):
                    errors.append(f"{raw_cell.coordinate} 公式没有可用缓存值")
    return errors


def _legacy_entry(
    *, snapshot: ParsedSnapshot, member: str, name: str, account_type: str,
    amount: object, location: str, institution: str | None = None,
    credit_limit: object = None, include: bool = True,
) -> None:
    amount_cents, warnings = _money_value(amount)
    limit_cents, limit_warnings = _money_value(credit_limit)
    warnings.extend(limit_warnings)
    institution = normalize_institution(institution)
    snapshot.entries.append(ParsedEntry(
        member_name=normalize_member_name(member),
        account_name=normalized_account_name(account_type, name, institution),
        account_type=account_type, amount_cents=amount_cents,
        credit_limit_cents=limit_cents, include_in_net_worth=include,
        institution=institution, source_sheet=snapshot.source_sheet,
        source_location=location, raw_name=name,
        raw_value="" if amount is None else str(amount), warnings=warnings,
    ))


def _sum_detail_values(values: list[object]) -> object:
    numbers = [round(value, 2) if isinstance(value, float) else value for value in values if isinstance(value, (int, float))]
    return round(sum(numbers), 2) if numbers else None


def _parse_legacy_sheet(
    title: str, rows: list[list[object]], raw_sheet, cached_sheet,
    date_overrides: dict[object, date] | None,
    known_debit_accounts: set[tuple[str, str]],
) -> ParsedSnapshot:
    try:
        workbook_date = _parse_date(rows[0][0])
    except (ValueError, IndexError):
        return ParsedSnapshot(
            snapshot_date=None, layout="legacy-family-monthly-matrix", source_sheet=title,
            status="blocked", blocking_errors=["A1 缺少有效盘点日期"],
        )
    period = _sheet_period(title, workbook_date)
    if period is None:
        return ParsedSnapshot(
            snapshot_date=None, source_date=workbook_date,
            layout="legacy-family-monthly-matrix", source_sheet=title,
            status="blocked", blocking_errors=["无法从工作表名称识别自然月"],
        )
    year, month = period
    override = _source_override(date_overrides, title, year, month)
    source_date = override or workbook_date
    snapshot = ParsedSnapshot(
        snapshot_date=date(year, month, 1), source_date=source_date,
        layout="legacy-family-monthly-matrix", source_sheet=title,
    )
    snapshot.snapshot_date = month_end(snapshot.snapshot_date)
    if workbook_date.year != year or workbook_date.month != month:
        if override is None or override.year != year or override.month != month:
            snapshot.status = "blocked"
            snapshot.blocking_errors.append(
                f"A1 日期 {workbook_date.isoformat()} 与工作表月份 {year:04d}-{month:02d} 冲突，且无 Markdown 佐证"
            )
        else:
            snapshot.warnings.append(
                f"A1 日期 {workbook_date.isoformat()} 与工作表月份冲突，采用 Markdown 日期 {override.isoformat()}"
            )
    snapshot.blocking_errors.extend(_formula_cache_errors(raw_sheet, cached_sheet))
    if snapshot.blocking_errors:
        snapshot.status = "blocked"

    headers = [str(value or "").strip() for value in rows[1]]
    total_index = next((index for index, value in enumerate(headers) if value == "总计"), len(headers))
    account_columns = list(range(1, total_index))

    for row_index in (2, 3):
        row = rows[row_index]
        label = str(_cell(row, 0) or "")
        member = "峰峰" if "大明" in label else "贤贤"
        for column in account_columns:
            institution = normalize_institution(headers[column])
            if not institution:
                continue
            value = _cell(row, column)
            if institution in {"微信", "支付宝"}:
                _legacy_entry(
                    snapshot=snapshot, member=member, name=institution,
                    account_type="wallet", amount=value,
                    location=f"{title}!{cached_sheet.cell(row_index + 1, column + 1).coordinate}",
                )
            else:
                if (member, institution) not in known_debit_accounts:
                    continue
                _legacy_entry(
                    snapshot=snapshot, member=member, name=institution,
                    account_type="debit_card", amount=value, institution=institution,
                    location=f"{title}!{cached_sheet.cell(row_index + 1, column + 1).coordinate}",
                )

    for limit_row, bill_row, member in ((4, 6, "峰峰"), (7, 9, "贤贤")):
        if bill_row >= len(rows):
            continue
        limits, bills = rows[limit_row], rows[bill_row]
        for column in account_columns:
            institution = normalize_institution(headers[column])
            limit, amount = _cell(limits, column), _cell(bills, column)
            if limit in (None, "") and amount in (None, ""):
                continue
            if not institution:
                continue
            if infer_account_type(institution) == "receivable":
                _legacy_entry(
                    snapshot=snapshot, member="家庭公共", name=institution,
                    account_type="receivable", amount=amount, include=False,
                    location=f"{title}!{cached_sheet.cell(bill_row + 1, column + 1).coordinate}",
                )
            else:
                _legacy_entry(
                    snapshot=snapshot, member=member, name=institution,
                    account_type="credit_card", amount=amount, institution=institution,
                    credit_limit=limit,
                    location=f"{title}!{cached_sheet.cell(bill_row + 1, column + 1).coordinate}",
                )

    if len(rows) > 10 and str(_cell(rows[10], 0) or "").strip() == "其他":
        other_row = rows[10]
        for column in range(1, min(total_index, len(other_row) - 1), 2):
            name = str(_cell(other_row, column) or "").strip()
            if not name:
                continue
            value = _cell(other_row, column + 1)
            normalized = name.lower()
            if "金峰证券" in name:
                member = "峰峰"
            elif "贤贤证券" in name:
                member = "贤贤"
            else:
                member = "家庭公共"
            account_type = "investment" if any(token in normalized for token in ("证券", "基金", "etf", "京东金融")) else infer_account_type(name)
            _legacy_entry(
                snapshot=snapshot, member=member, name=name, account_type=account_type,
                amount=value, location=f"{title}!{cached_sheet.cell(11, column + 2).coordinate}",
            )
        source_total = _cell(other_row, total_index)
        if isinstance(source_total, (int, float)):
            snapshot.legacy_summary["investment_source_total_cents"] = _money_value(source_total)[0]

    for row_index in range(11, len(rows)):
        row = rows[row_index]
        label = str(_cell(row, 0) or "").strip()
        if not label:
            continue
        summary_key = SUMMARY_ROWS.get(label)
        if summary_key:
            value = _cell(row, total_index)
            amount, warnings = _money_value(value)
            snapshot.legacy_summary[summary_key] = amount
            snapshot.warnings.extend(warnings)
            continue
        account_type = infer_account_type(label)
        if account_type not in {"receivable", "other_liability"}:
            continue
        detail_amount = _sum_detail_values([_cell(row, column) for column in range(1, total_index)])
        _legacy_entry(
            snapshot=snapshot, member="家庭公共", name=label,
            account_type=account_type, amount=detail_amount,
            include="不计入总数" not in label,
            location=f"{title}!第 {row_index + 1} 行",
        )
        row_total = _cell(row, total_index)
        if isinstance(row_total, (int, float)):
            total_cents = _money_value(row_total)[0]
            snapshot.legacy_summary.setdefault("ignored_auxiliary_totals_cents", 0)
            snapshot.legacy_summary["ignored_auxiliary_totals_cents"] += total_cents or 0
        for column in range(1, total_index):
            detail_value = _cell(row, column)
            if isinstance(detail_value, (int, float)):
                snapshot.legacy_summary[
                    f"auxiliary_detail_{row_index + 1}_{column + 1}_cents"
                ] = _money_value(detail_value)[0]

    seen: set[tuple[str, str, str, str | None]] = set()
    for entry in snapshot.entries:
        key = (entry.member_name, entry.account_type, entry.account_name, entry.institution)
        if key in seen:
            snapshot.blocking_errors.append(f"重复明细：{entry.member_name}/{entry.account_name}")
        seen.add(key)
    if snapshot.blocking_errors:
        snapshot.status = "blocked"
    return snapshot


def _known_debit_accounts(workbook) -> set[tuple[str, str]]:
    known: set[tuple[str, str]] = set()
    for sheet in workbook.worksheets:
        if sheet.title.strip() == "工作表1":
            continue
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not _legacy_layout(rows):
            continue
        headers = [str(value or "").strip() for value in rows[1]]
        total_index = next((index for index, value in enumerate(headers) if value == "总计"), len(headers))
        for row_index in (2, 3):
            member = "峰峰" if "大明" in str(_cell(rows[row_index], 0) or "") else "贤贤"
            for column in range(1, total_index):
                institution = normalize_institution(headers[column])
                if institution and institution not in {"微信", "支付宝"} and _cell(rows[row_index], column) not in (None, ""):
                    known.add((member, institution))
    return known


def parse_excel(
    content: bytes, date_overrides: dict[object, date] | None = None,
) -> list[ParsedSnapshot]:
    cached_workbook = raw_workbook = None
    results: list[ParsedSnapshot] = []
    try:
        cached_workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        raw_workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=False)
        known_debit_accounts = _known_debit_accounts(cached_workbook)
        for cached_sheet in cached_workbook.worksheets:
            title = cached_sheet.title
            raw_sheet = raw_workbook[title]
            rows = [list(row) for row in cached_sheet.iter_rows(values_only=True)]
            if title.strip() == "工作表1":
                results.append(ParsedSnapshot(
                    snapshot_date=None, layout="legacy-helper", source_sheet=title,
                    status="ignored", warnings=["重复辅助页，已忽略"],
                ))
                continue
            if not rows or not any(any(value not in (None, "") for value in row) for row in rows):
                results.append(ParsedSnapshot(
                    snapshot_date=None, layout="empty", source_sheet=title,
                    status="ignored", warnings=["空工作表，已忽略"],
                ))
                continue
            if _legacy_layout(rows):
                results.append(_parse_legacy_sheet(
                    title, rows, raw_sheet, cached_sheet, date_overrides, known_debit_accounts,
                ))
                continue
            try:
                parsed = rows_to_snapshots(rows, title)
            except ValueError as exc:
                results.append(ParsedSnapshot(
                    snapshot_date=None, layout="unknown", source_sheet=title,
                    status="blocked", blocking_errors=[str(exc)],
                ))
            else:
                results.extend(parsed)
    except (BadZipFile, InvalidFileException, KeyError, OSError, EOFError, ParseError, ValueError) as exc:
        raise ValueError("Excel 文件损坏或不是有效的 XLSX/XLSM 工作簿") from exc
    finally:
        if cached_workbook is not None:
            cached_workbook.close()
        if raw_workbook is not None:
            raw_workbook.close()
    if not results:
        raise ValueError("Excel 工作簿中没有可解析的工作表")
    return results
