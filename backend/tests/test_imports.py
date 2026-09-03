import io
from datetime import date

from openpyxl import Workbook
import pytest
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.database import Base, create_sqlite_engine
from app.importers.common import ParsedEntry, ParsedSnapshot
from app.importers.legacy_markdown import parse_legacy_markdown
from app.importers.tabular import parse_csv, parse_excel
from app.models import ImportRecord, Snapshot
from app.services.imports import import_snapshots


LEGACY_MARKDOWN = """
## 2025年12月25日

### 一、大明同学明细

#### 钱包 / 支付平台
微信：987.63
支付宝：

#### 信用卡
招商银行信用卡：额度 64000，待还 -18.52

#### 应收款
Momo 借款：11704

### 三、总计
总资产：12691.63
总负债：-18.52
顺差：12710.15
"""


def excel_bytes(*rows: list[object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_markdown_parser_preserves_null_negative_and_receivable():
    snapshots = parse_legacy_markdown(LEGACY_MARKDOWN)
    assert len(snapshots) == 1
    entries = {entry.account_name: entry for entry in snapshots[0].entries}
    assert entries["微信"].amount_cents == 98_763
    assert entries["支付宝"].amount_cents is None
    assert entries["招商银行信用卡"].amount_cents == -1_852
    assert entries["招商银行信用卡"].credit_limit_cents == 6_400_000
    assert entries["Momo 借款"].account_type == "receivable"


def test_markdown_import_creates_snapshot_and_report(client):
    response = client.post(
        "/api/import/legacy",
        files={"file": ("history.md", LEGACY_MARKDOWN.encode("utf-8"), "text/markdown")},
    )
    assert response.status_code == 200, response.text
    report = response.json()
    assert report["success_rows"] == 4
    snapshots = client.get("/api/snapshots?status=completed").json()
    assert len(snapshots) == 1
    detail = client.get(f"/api/snapshots/{snapshots[0]['id']}").json()
    assert any(entry["legacy_raw_value"] for entry in detail["entries"])


def test_markdown_preview_is_read_only_and_supports_gb18030(client):
    response = client.post(
        "/api/import/legacy/preview",
        files={"file": ("history.md", LEGACY_MARKDOWN.encode("gb18030"), "text/markdown")},
    )
    assert response.status_code == 200, response.text
    preview = response.json()
    assert preview["source_type"] == "markdown"
    assert preview["detected_encoding"] == "gb18030"
    assert preview["total_snapshots"] == 1
    assert preview["total_rows"] == 4
    assert preview["importable_rows"] == 4
    assert client.get("/api/imports").json() == []
    assert client.get("/api/snapshots").json() == []


def test_csv_parser_supports_blank_and_negative():
    content = "盘点日期,家庭成员,账户名称,账户类型,金额\n2026-01-31,大明,微信,wallet,12.35\n2026-01-31,大明,信用卡,credit_card,-18.52\n2026-01-31,大明,空账户,other_asset,\n"
    snapshots = parse_csv(content.encode("utf-8"))
    assert [item.amount_cents for item in snapshots[0].entries] == [1_235, -1_852, None]


def test_excel_parser_uses_same_tabular_contract():
    content = excel_bytes(
        ["盘点日期", "家庭成员", "账户名称", "账户类型", "金额"],
        ["2026-02-28", "贤贤", "投资账户", "investment", "1000.01"],
    )
    snapshots = parse_excel(content)
    assert snapshots[0].entries[0].amount_cents == 100_001


def test_csv_preview_and_import_endpoints_support_gb18030(client):
    content = (
        "盘点日期,家庭成员,账户名称,账户类型,金额\n"
        "2026-01-31,大明,微信,wallet,12.35\n"
    ).encode("gb18030")
    preview_response = client.post(
        "/api/import/tabular/preview",
        files={"file": ("history.csv", content, "text/csv")},
    )
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    assert preview["detected_encoding"] == "gb18030"
    assert preview["total_rows"] == 1
    assert client.get("/api/imports").json() == []

    import_response = client.post(
        "/api/import/tabular",
        files={"file": ("history.csv", content, "text/csv")},
    )
    assert import_response.status_code == 200, import_response.text
    assert import_response.json()["success_rows"] == 1


def test_excel_preview_is_read_only_then_imports_and_reports_duplicate(client):
    content = excel_bytes(
        ["盘点日期", "家庭成员", "账户名称", "账户类型", "金额"],
        ["2026-02-28", "贤贤", "投资账户", "investment", "1000.01"],
    )
    files = {"file": ("history.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    preview_response = client.post("/api/import/tabular/preview", files=files)
    assert preview_response.status_code == 200, preview_response.text
    assert preview_response.json()["importable_rows"] == 1
    assert client.get("/api/imports").json() == []

    import_response = client.post("/api/import/tabular", files=files)
    assert import_response.status_code == 200, import_response.text
    assert import_response.json()["success_rows"] == 1

    duplicate_response = client.post("/api/import/tabular/preview", files=files)
    assert duplicate_response.status_code == 200, duplicate_response.text
    duplicate = duplicate_response.json()
    assert duplicate["duplicate_snapshots"] == 1
    assert duplicate["importable_rows"] == 0
    assert duplicate["snapshots"][0]["will_skip"] is True


def test_import_preview_rejects_invalid_or_unsupported_files(client):
    broken_excel = client.post(
        "/api/import/tabular/preview",
        files={"file": ("broken.xlsx", b"not-an-excel-file", "application/octet-stream")},
    )
    assert broken_excel.status_code == 422
    assert "Excel 文件损坏" in broken_excel.json()["detail"]

    empty_rows = "盘点日期,家庭成员,账户名称,金额\n无效日期,,,\n".encode("utf-8")
    no_valid_rows = client.post(
        "/api/import/tabular/preview",
        files={"file": ("empty.csv", empty_rows, "text/csv")},
    )
    assert no_valid_rows.status_code == 422
    assert "没有可导入的有效数据行" in no_valid_rows.json()["detail"]

    unsupported = client.post(
        "/api/import/tabular/preview",
        files={"file": ("legacy.xls", b"old-excel", "application/vnd.ms-excel")},
    )
    assert unsupported.status_code == 422
    assert "不支持旧版 XLS" in unsupported.json()["detail"]


def test_import_preview_enforces_upload_limit(client):
    response = client.post(
        "/api/import/legacy/preview",
        files={"file": ("large.md", b"x" * (20 * 1024 * 1024 + 1), "text/markdown")},
    )
    assert response.status_code == 413


def test_exported_amount_columns_do_not_multiply_cents_by_one_hundred():
    content = "snapshot_date,member_name,account_name,account_type,amount,amount_cents\n2026-03-31,大明,微信,wallet,12.35,1235\n"
    snapshots = parse_csv(content.encode("utf-8"))
    assert snapshots[0].entries[0].amount_cents == 1_235


def test_cents_only_column_is_read_as_cents():
    content = "snapshot_date,member_name,account_name,account_type,amount_cents\n2026-03-31,大明,微信,wallet,1235\n"
    snapshots = parse_csv(content.encode("utf-8"))
    assert snapshots[0].entries[0].amount_cents == 1_235


HORIZONTAL_MARKDOWN = """
## 2026年8月19日
### 一、大明同学明细
| 微信 | 支付宝 |
| --- | --- |
| 12.30 | 0 |
| 储蓄卡 | 名称 | 余额 |
| --- | --- | --- |
| | 招商银行 | 100.01 |
| 小记 | | 100.01 |
| 信用卡 | 名称 | 额度 | 需还款 |
| --- | --- | --- | --- |
| 11号 | 招商银行 | 64000 | -18.52 |
| | 借款待收回（不计入总数） | | 900 |
| 小记 | | | 881.48 |
### 三、总计
| 家庭总余额 | 112.31 |
| --- | --- |
| 家庭总负债 | -18.52 |
| 顺差 | 130.83 |
"""


def test_horizontal_markdown_uses_table_semantics_and_normalizes_names():
    snapshot = parse_legacy_markdown(HORIZONTAL_MARKDOWN)[0]
    assert snapshot.source_date.isoformat() == "2026-08-19"
    assert snapshot.snapshot_date.isoformat() == "2026-08-31"
    entries = {(entry.member_name, entry.account_name): entry for entry in snapshot.entries}
    assert entries[("峰峰", "微信")].amount_cents == 1_230
    assert entries[("峰峰", "支付宝")].amount_cents == 0
    assert entries[("峰峰", "招商银行储蓄卡")].amount_cents == 10_001
    card = entries[("峰峰", "招商银行信用卡")]
    assert card.amount_cents == -1_852
    assert card.credit_limit_cents == 6_400_000
    assert card.billing_day == 11
    receivable = entries[("家庭公共", "借款待收回（不计入总数）")]
    assert receivable.account_type == "receivable"
    assert receivable.include_in_net_worth is False
    assert not ({"小记", "家庭总余额", "11号"} & {entry.account_name for entry in snapshot.entries})


def legacy_matrix_workbook(*, conflict: bool = False, missing_formula_cache: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "家庭存款明细(26.4月）"
    sheet.append(["2026-04-26"])
    sheet.append([None, "招商银行", "平安银行", "微信", "支付宝", "总计"])
    sheet.append(["大明同学存款", 100, None, 12.3, 0, "=SUM(B3:E3)" if missing_formula_cache else 112.3])
    sheet.append(["贤贤存款", 200, 3, 20, 5, 228])
    sheet.append(["大明同学信用卡额度", 64_000, None, None, None, None])
    sheet.append(["大明同学可使用额度", 63_990, None, None, None, None])
    sheet.append(["大明同学信用卡账单", 10, None, None, None, 10])
    sheet.append(["贤贤信用卡额度", 53_000, None, None, None, None])
    sheet.append(["贤贤可使用额度", 52_980, None, None, None, None])
    sheet.append(["贤贤信用卡账单", 20, None, None, None, 20])
    sheet.append(["其他", "黄金ETF", 300, "基金", 400, 700])
    if conflict:
        sheet.title = "家庭存款明细(11月）"
        sheet["A1"] = "2024-08-01"
    helper = workbook.create_sheet("工作表1")
    helper.append(["重复辅助页"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_legacy_excel_lists_each_sheet_and_ignores_helper():
    snapshots = parse_excel(legacy_matrix_workbook())
    assert [item.status for item in snapshots] == ["importable", "ignored"]
    matrix = snapshots[0]
    assert matrix.layout == "legacy-family-monthly-matrix"
    assert matrix.snapshot_date.isoformat() == "2026-04-30"
    entries = {(entry.member_name, entry.account_name): entry for entry in matrix.entries}
    assert entries[("峰峰", "招商银行储蓄卡")].amount_cents == 10_000
    assert entries[("峰峰", "支付宝")].amount_cents == 0
    assert entries[("峰峰", "招商银行信用卡")].credit_limit_cents == 6_400_000
    assert entries[("家庭公共", "黄金ETF")].account_type == "investment"


def test_legacy_excel_blocks_month_conflict_without_markdown_evidence():
    snapshots = parse_excel(legacy_matrix_workbook(conflict=True))
    blocked = snapshots[0]
    assert blocked.status == "blocked"
    assert "日期" in blocked.blocking_errors[0]
    assert "冲突" in blocked.blocking_errors[0]


def test_legacy_excel_blocks_formula_without_cached_value():
    snapshots = parse_excel(legacy_matrix_workbook(missing_formula_cache=True))
    assert snapshots[0].status == "blocked"
    assert any("公式没有可用缓存值" in error for error in snapshots[0].blocking_errors)


def test_import_deduplicates_completed_snapshots_by_natural_month(client):
    first = "盘点日期,家庭成员,账户名称,账户类型,金额\n2026-04-01,峰峰,微信,wallet,1\n"
    second = "盘点日期,家庭成员,账户名称,账户类型,金额\n2026-04-29,峰峰,微信,wallet,2\n"
    imported = client.post(
        "/api/import/tabular",
        files={"file": ("first.csv", first.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    preview = client.post(
        "/api/import/tabular/preview",
        files={"file": ("second.csv", second.encode(), "text/csv")},
    ).json()
    assert preview["snapshots"][0]["status"] == "duplicate"
    assert preview["importable_rows"] == 0
    stored = client.get("/api/snapshots?status=completed").json()
    assert stored[0]["snapshot_date"] == "2026-04-30"


def test_preview_blocks_duplicate_account_inside_one_snapshot(client):
    content = (
        "盘点日期,家庭成员,账户名称,账户类型,金额\n"
        "2026-04-01,峰峰,微信,wallet,1\n"
        "2026-04-20,峰峰,微信,wallet,2\n"
    ).encode()
    preview = client.post(
        "/api/import/tabular/preview",
        files={"file": ("duplicate.csv", content, "text/csv")},
    ).json()
    assert preview["snapshots"][0]["status"] == "blocked"
    assert any("重复账户" in error for error in preview["snapshots"][0]["blocking_errors"])
    assert client.get("/api/snapshots").json() == []


def test_commit_false_can_be_rolled_back_as_one_transaction(tmp_path):
    engine = create_sqlite_engine(tmp_path / "rollback.db")
    Base.metadata.create_all(engine)
    parsed = ParsedSnapshot(
        snapshot_date=date(2026, 4, 30),
        entries=[ParsedEntry("峰峰", "微信", "wallet", 100)],
    )
    with pytest.raises(RuntimeError):
        with Session(engine) as session:
            with session.begin():
                import_snapshots(session, [parsed], "rollback.csv", "csv", commit=False)
                raise RuntimeError("验证整批回滚")
    with Session(engine) as session:
        assert session.scalar(select(func.count()).select_from(Snapshot)) == 0
        assert session.scalar(select(func.count()).select_from(ImportRecord)) == 0
    engine.dispose()
